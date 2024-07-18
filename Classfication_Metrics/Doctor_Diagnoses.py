#!/usr/bin/env python3

import openpyxl
from openpyxl import Workbook
import random

# Function to calculate binary classification metrics
def calculate_metrics(TP, TN, FP, FN):
	accuracy = (TP + TN) / (TP + TN + FP + FN)
	precision = TP / (TP + FP) if (TP + FP) != 0 else 0
	recall = TP / (TP + FN) if (TP + FN) != 0 else 0
	f1_score = (2 * precision * recall) / (precision + recall) if (precision + recall) != 0 else 0
	return accuracy, precision, recall, f1_score

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Doctor Diagnoses"

# Add headers to the columns
ws.append(['Scenario', 'Count', 'Description'])
ws.append(['True Positive (TP)', 0, 'Doctor correctly tells Rachel she is pregnant'])
ws.append(['True Negative (TN)', 0, 'Doctor correctly tells Ross he is not pregnant'])
ws.append(['False Positive (FP)', 0, 'Doctor incorrectly tells Ross he is pregnant when he isn\'t'])
ws.append(['False Negative (FN)', 0, 'Doctor incorrectly tells Rachel she is not pregnant when she is'])
ws.append([])
ws.append(['True Positive', 'True Negative', 'False Positive', 'False Negative', 'Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC-ROC'])

# Generate dummy data
TP = random.randint(50, 100)
TN = random.randint(50, 100)
FP = random.randint(0, 50)
FN = random.randint(0, 50)

# Update the counts in the scenarios section
ws['B2'] = TP
ws['B3'] = TN
ws['B4'] = FP
ws['B5'] = FN

# Add the data to the sheet
ws.append([TP, TN, FP, FN])

# Add formulas for each metric (linked to cell references)
ws['E8'] = "=(A8 + B8) / (A8 + B8 + C8 + D8)"
ws['F8'] = "=A8 / (A8 + C8)"
ws['G8'] = "=A8 / (A8 + D8)"
ws['H8'] = "=(2 * F8 * G8) / (F8 + G8)"
ws['I8'] = f"={random.uniform(0.5, 1.0)}"

# Save the workbook
wb.save('Doctor_Diagnoses.xlsx')
