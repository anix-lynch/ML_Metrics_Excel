import openpyxl
from openpyxl import Workbook
import random

# Function to calculate binary classification metrics
def calculate_metrics(TP, TN, FP, FN):
    accuracy = (TP + TN) / (TP + TN + FP + FN)
    precision = TP / (TP + FP) if (TP + FP) != 0 else 0
    recall = TP / (TP + FN) if (TP + FN) != 0 else 0
    f1_score = (2 * precision * recall) / (precision + recall) if (precision + recall) != 0 else 0
    return accuracy, precision, recall, f1_score

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Classification Metrics"

# Add headers to the columns
ws.append(['True Positive', 'True Negative', 'False Positive', 'False Negative', 'Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC-ROC'])

# Generate dummy data
TP = random.randint(50, 100)
TN = random.randint(50, 100)
FP = random.randint(0, 50)
FN = random.randint(0, 50)

# Add the data to the sheet
ws.append([TP, TN, FP, FN])

# Add formulas for each metric (linked to cell references)
ws['E2'] = "=(A2 + B2) / (A2 + B2 + C2 + D2)"
ws['F2'] = "=A2 / (A2 + C2)"
ws['G2'] = "=A2 / (A2 + D2)"
ws['H2'] = "=(2 * F2 * G2) / (F2 + G2)"
ws['I2'] = f"={random.uniform(0.5, 1.0)}"

# Save the workbook
wb.save('Classification_Metrics.xlsx')
