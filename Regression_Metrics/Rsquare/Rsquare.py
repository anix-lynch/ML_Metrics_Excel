#!/usr/bin/env python3

import openpyxl
from openpyxl import Workbook
import random

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Add headers to the columns
ws['A1'] = 'Actual Values'
ws['B1'] = 'Predicted Values'
ws['C1'] = 'Absolute Error'
ws['D1'] = 'MAE'
ws['E1'] = 'Squared Error'
ws['F1'] = 'R²'

# Generate dummy data
for row in range(2, 11):
	actual_value = random.uniform(10, 100)
	predicted_value = actual_value + random.uniform(-10, 10)
	ws[f'A{row}'] = actual_value
	ws[f'B{row}'] = predicted_value
	ws[f'C{row}'] = f'=ABS(A{row} - B{row})'
	ws[f'E{row}'] = f'=(A{row} - B{row})^2'
	
# Calculate the mean absolute error (MAE)
ws['D2'] = '=AVERAGE(C2:C10)'

# Calculate the total sum of squares (SS_tot)
for row in range(2, 11):
	ws[f'G{row}'] = f'=(A{row} - AVERAGE(A2:A10))^2'
ws['H2'] = '=SUM(G2:G10)'

# Calculate the residual sum of squares (SS_res)
ws['I2'] = '=SUM(E2:E10)'

# Calculate R-squared (R²)
ws['F2'] = '=1 - (I2 / H2)'

# Save the workbook
wb.save('mae_r2_example.xlsx')
