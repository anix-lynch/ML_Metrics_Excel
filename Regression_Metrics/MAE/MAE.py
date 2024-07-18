#!/usr/bin/env python3

import openpyxl
from openpyxl import Workbook
import random

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Add headers to the columns
ws['A1'] = 'Actual Values'
ws['B1'] = 'Predicted Values'
ws['C1'] = 'Absolute Error'
ws['D1'] = 'MAE'

# Generate dummy data
for row in range(2, 11):
	actual_value = random.uniform(10, 100)
	predicted_value = actual_value + random.uniform(-10, 10)
	ws[f'A{row}'] = actual_value
	ws[f'B{row}'] = predicted_value
	ws[f'C{row}'] = f'=ABS(A{row} - B{row})'
	
# Calculate the mean absolute error (MAE)
ws['D2'] = '=AVERAGE(C2:C10)'

# Save the workbook
wb.save('mae_example.xlsx')
